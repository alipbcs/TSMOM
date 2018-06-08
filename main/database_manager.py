import numpy as np
import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import datetime
from os import path
from typing import Tuple, List, Optional, Union

contract_dict = {'FTSE 100 IDX FUT': ['Equity', 'Developed'],
                 'SWISS MKT IX FUTR': ['Equity', 'Developed'],
                 'E-Mini Russ 2000': ['Equity', 'Developed'],
                 'S&P/TSX FIN IX FU': ['Equity', 'Developed'],
                 'EURO STOXX 50': ['Equity', 'Developed'],
                 'S&P 500 FUTURE': ['Equity', 'Developed'],
                 'NIKKEI 225 (OSE)': ['Equity', 'Developed'],
                 'CAC40 10 EURO FUT': ['Equity', 'Developed'],
                 'DAX INDEX FUTURE': ['Equity', 'Developed'],
                 'HANG SENG IDX FUT': ['Equity', 'Developed'],
                 'NASDAQ 100 E-MINI': ['Equity', 'Developed'],
                 'IBEX 35 INDX FUTR': ['Equity', 'Developed'],
                 'S&P/TSX 60 IX FUT': ['Equity', 'Developed'],
                 'AMSTERDAM IDX FUT': ['Equity', 'Developed'],
                 'DJIA MINI e-CBOT': ['Equity', 'Developed'],
                 'FTSE/MIB IDX FUT': ['Equity', 'Emerging'],
                 'BOVESPA INDEX FUT': ['Equity', 'Emerging'],
                 'FTSE-TWSE 50 FUT': ['Equity', 'Emerging'],
                 'SPI 200 FUTURES': ['Equity', 'Emerging'],
                 'SGX Nifty 50': ['Equity', 'Emerging'],
                 'FTSE KLCI FUTURE': ['Equity', 'Emerging'],
                 'MICEX Index Futur': ['Equity', 'Emerging'],
                 '90DAY EURO$ FUTR': ['Interest Rates'],
                 'FED FUND 30DAY': ['Interest Rates'],
                 'EURO-BUND FUTURE': ['Interest Rates'],
                 'Euro-BTP Future': ['Interest Rates'],
                 'US 10YR NOTE': ['Interest Rates'],
                 'US LONG BOND(CBT)': ['Interest Rates'],
                 'Euro-OAT Future': ['Interest Rates'],
                 'CAN 10YR BOND FUT': ['Interest Rates'],
                 'JPN 10Y BOND(OSE)': ['Interest Rates'],
                 'SWISS FED BND FUT': ['Interest Rates'],
                 'EURO-SCHATZ FUT': ['Interest Rates'],
                 'EURO BUXL 30Y BND': ['Interest Rates'],
                 'US 5YR NOTE (CBT)': ['Interest Rates'],
                 'US 2YR NOTE (CBT)': ['Interest Rates'],
                 'LONG GILT FUTURE': ['Interest Rates'],
                 'EURO-BOBL FUTURE': ['Interest Rates'],
                 '90DAY STERLING FU': ['Interest Rates'],
                 '3MO EURO EURIBOR': ['Interest Rates'],
                 'MEXICAN PESO FUT': ['Currency'],
                 'CHF CURRENCY FUT': ['Currency'],
                 'JPN YEN CURR FUT': ['Currency'],
                 'BP CURRENCY FUT': ['Currency'],
                 'AUDUSD Crncy Fut': ['Currency'],
                 'NEW ZEALAND $ FUT': ['Currency'],
                 'C$ CURRENCY FUT': ['Currency'],
                 'EURO FX CURR FUT': ['Currency'],
                 'KC HRW WHEAT FUT': ['Commodity', 'Grains'],
                 'WHEAT FUTURE(CBT)': ['Commodity', 'Grains'],
                 'CORN FUTURE': ['Commodity', 'Grains'],
                 'SOYBEAN OIL FUTR': ['Commodity', 'Grains'],
                 'SOYBEAN FUTURE': ['Commodity', 'Grains'],
                 'SOYBEAN MEAL FUTR': ['Commodity', 'Grains'],
                 'OAT FUTURE': ['Commodity', 'Grains'],
                 'ROUGH RICE (CBOT)': ['Commodity', 'Grains'],
                 'LUMBER FUTURE': ['Commodity', 'Softs'],
                 'COTTON NO.2 FUTR': ['Commodity', 'Softs'],
                 '''COFFEE 'C' FUTURE''': ['Commodity', 'Softs'],
                 'COCOA FUTURE': ['Commodity', 'Softs'],
                 'FCOJ-A FUTURE': ['Commodity', 'Softs'],
                 'SUGAR #11 (WORLD)': ['Commodity', 'Softs'],
                 'MILK FUTURE': ['Commodity', 'Softs'],
                 'BRENT CRUDE FUTR': ['Commodity', 'Energy'],
                 'GASOLINE RBOB FUT': ['Commodity', 'Energy'],
                 'WTI CRUDE FUTURE': ['Commodity', 'Energy'],
                 'NATURAL GAS FUTR': ['Commodity', 'Energy'],
                 'NY Harb ULSD Fut': ['Commodity', 'Energy'],
                 'ICE RTD CAL YEAR': ['Commodity', 'Energy'],
                 'SILVER FUTURE': ['Commodity', 'Precious Metal'],
                 'PALLADIUM FUTURE': ['Commodity', 'Precious Metal'],
                 'PLATINUM FUTURE': ['Commodity', 'Precious Metal'],
                 'GOLD 100 OZ FUTR': ['Commodity', 'Precious Metal'],
                 'COPPER FUTURE': ['Commodity', 'Metal'],
                 'LME PRI ALUM FUTR': ['Commodity', 'Metal'],
                 'LME ZINC FUTURE': ['Commodity', 'Metal'],
                 'LME LEAD FUTURE': ['Commodity', 'Metal'],
                 'CATTLE FEEDER FUT': ['Commodity', 'Meat'],
                 'LIVE CATTLE FUTR': ['Commodity', 'Meat'],
                 'LEAN HOGS FUTURE': ['Commodity', 'Meat']}

ROOT_DIR = path.dirname(path.abspath(__file__ + "/.."))


class DatabaseManager(object):
    """
    Database manager for Stevens Continuous Futures (SCF) & Bloomberg datasets.
    """

    # Borg design pattern
    __shared_state = {}

    def __init__(self):
        """
        Initializes a new DatabaseManager instance and opens a connection to
        sqlite3 database'SCF' in the same folder. (A new one is created if it
        doesn't already exits.)
        """
        self.__dict__ = self.__shared_state

        if not hasattr(self, 'con'):
            print(ROOT_DIR)
            self.con = sqlite3.connect(ROOT_DIR + '/data/SCF.db')

        if not hasattr(self, 'quandl_dataset_names'):
            self.quandl_dataset_names = []

            cursor = self.con.execute('''SELECT tbl_name FROM sqlite_master 
                                        WHERE type='table' AND 
                                        tbl_name LIKE 'quandl%' ''')

            result = cursor.fetchall()

            for tbl in result:
                self.quandl_dataset_names.append(tbl[0])

        if not hasattr(self, 'bloom_dataset_names'):
            self.bloom_dataset_names = []

            cursor = self.con.execute('''SELECT tbl_name FROM sqlite_master 
                                        WHERE type='table' AND 
                                        tbl_name LIKE 'bloom%' ''')

            result = cursor.fetchall()

            for tbl in result:
                self.bloom_dataset_names.append(tbl[0])

            self.bloom_dataset_names.remove('bloom_info')

        if not hasattr(self, 'bloom_column_names'):
            self.bloom_column_names = ['Dates', 'PX_LAST', 'PX_OPEN', 'PX_HIGH', 'PX_LOW']

        if not hasattr(self, 'bloom_to_qunadl_dict'):
            self.bloom_to_qunadl_dict = {}
            self.quandl_to_bloom_dict = {}

            self.__build_quandl_to_bloom_dict()

        if not hasattr(self, 'type_to_table_dict'):
            self.type_to_table_dict = {}
            self.subtype_to_table_dict = {}

            self.__build_type_to_table_dict()

    def __build_quandl_to_bloom_dict(self) -> None:
        """
        Builds a dictionary containing mappings from quandl datasets to bloom datasets.
        """
        wb = load_workbook(ROOT_DIR + '/data/quandl/quandl_Bloom_dict.xlsx', data_only=True)
        sh = wb['quandl_Bloom_dict']

        for i in range(2, 80):
            quandl_symbol = sh['a' + str(i)].value
            bloom_contract_name = sh['n' + str(i)].value

            if bloom_contract_name is not None:
                result = self.con.execute('''SELECT tbl_name FROM sqlite_master 
                                          WHERE type='table' and 
                                          tbl_name LIKE '{}' ESCAPE '\\' '''.format(
                                                'quandl%\_' + quandl_symbol + '1_OR')
                )

                tbl_quandl = result.fetchone()

                if tbl_quandl is None:
                    continue

                cursor = self.con.execute('''SELECT table_name
                                          FROM bloom_info
                                          WHERE contract_name='{}' '''.format(
                    bloom_contract_name)
                )

                tbl_bloom = cursor.fetchone()

                if tbl_bloom is None:
                    continue

                if self.bloom_to_qunadl_dict.get(tbl_bloom[0]) is None:
                    self.bloom_to_qunadl_dict[tbl_bloom[0]] = []

                if self.quandl_to_bloom_dict.get(tbl_quandl[0]) is None:
                    self.quandl_to_bloom_dict[tbl_quandl[0]] = []

                self.quandl_to_bloom_dict[tbl_quandl[0]].append(tbl_bloom[0])
                self.bloom_to_qunadl_dict[tbl_bloom[0]].append(tbl_quandl[0])

    def __build_type_to_table_dict(self) -> None:
        """
        Builds a dictionary containing mappings from type to table name.
        """
        for tbl in self.bloom_dataset_names:
            info = self.get_info(tbl)

            if info is not None:
                if info[3] is not None:
                    if self.type_to_table_dict.get(info[3]) is None:
                        self.type_to_table_dict[info[3]] = []

                    if self.bloom_to_qunadl_dict.get(tbl) is not None:
                        self.type_to_table_dict[info[3]].extend(self.bloom_to_qunadl_dict[tbl])

                    self.type_to_table_dict[info[3]].append(tbl)

                if info[4] is not None:
                    if self.subtype_to_table_dict.get(info[4]) is None:
                        self.subtype_to_table_dict[info[4]] = []

                    if self.bloom_to_qunadl_dict.get(tbl) is not None:
                        self.subtype_to_table_dict[info[4]].extend(self.bloom_to_qunadl_dict[tbl])

                    self.subtype_to_table_dict[info[4]].append(tbl)

    def import_bloom_from_xlsx(self, verbose: bool = False) -> None:
        """
        Imports Excel datasets, (assumed to be in 'bloom' folder relative to current
        working directory) to sqlite3 database SCF.db in current working directory.
        :param verbose: print dataset names during construction
        """
        for excel_dataset in Path(ROOT_DIR + '/bloom').glob('*.xlsx'):
            if verbose:
                print(excel_dataset.name)

            table_name = 'bloom_' + excel_dataset.name.split('.')[0]
            wb = load_workbook('./bloom/' + excel_dataset.name, data_only=True)
            sh = wb['Sheet1']
            contract_full_name = sh["G2"].value
            symbol = sh['c1'].value
            name_pieces = str.split(contract_full_name)
            contract_full = ' '.join(name_pieces[:-1])

            df = pd.read_excel('./bloom/' + excel_dataset.name, sheet_name='Sheet1', skiprows=4,
                               usecols=[0, 1, 2, 3, 4])

            df['Dates'] = df['Dates'].astype(np.str_)

            contract_type = contract_dict.get(contract_full)

            contract_has_type, contract_has_subtype = False, False

            if contract_type is not None:
                contract_has_type = True

                if len(contract_type) > 1:
                    contract_has_subtype = True

            contract_full = contract_full.replace("'", '')

            if contract_has_subtype:
                self.con.execute('''INSERT INTO bloom_info values('{}', '{}',
                                                                  '{}', '{}',  
                                                                  '{}')'''.format(table_name, contract_full, symbol,
                                                                                  contract_type[0], contract_type[1]))
            elif contract_has_type:
                self.con.execute('''INSERT INTO bloom_info values('{}', '{}', 
                                                                  '{}', '{}',
                                                                  null
                                                                  )'''.format(table_name, contract_full, symbol,
                                                                              contract_type[0]))
            else:
                self.con.execute('''INSERT INTO bloom_info values('{}', '{}',
                                                                  '{}', null,
                                                                  null
                                                                  )'''.format(table_name, contract_full, symbol))

            self.con.execute('''CREATE TABLE ''' + table_name + '''( 
                                Dates date not null,             
                                PX_LAST REAL,                      
                                PX_OPEN REAL,                     
                                PX_HIGH REAL,                                    
                                PX_LOW REAL,
                                PRIMARY KEY(Dates)
                                )'''
                             )

            df.to_sql(name=table_name, con=self.con, if_exists='append',
                      index=False)

        self.con.commit()

    def import_quandl_from_csv(self) -> None:
        """
        Imports .csv datasets, (assumed to be in 'SCF_csv' folder relative to current
        working directory) to sqlite3 database SCF.db in current working directory.
        """
        column_names = ['Date', 'Open', 'High', 'Low', 'Settle', 'Volume',
                        'Prev_day_OI']

        for df_path in Path(ROOT_DIR + '/SCF_csv').glob('*.csv'):
            table_name = df_path.name.split('.')[0]

            self.con.execute('''CREATE TABLE ''' + table_name + '''( 
                            Date date NOT NULL,             
                            Open REAL,                      
                            High REAL,                     
                            Low REAL,                
                            Settle REAL,
                            Volume REAL,
                            Prev_day_OI REAL,                               
                            primary key(Date)                                       
                            )'''
                             )

            df_temp = pd.read_csv('./SCF_csv/' + df_path.name, names=column_names)
            df_temp.to_sql(name=table_name, con=self.con, if_exists='append', index=False)

        self.con.commit()

    def get_table(self, tbl_name: str) -> Union[Tuple[pd.DataFrame, Tuple[str, str, str, str, str]], Tuple[None, None]]:
        """
        Retrieves table with accompanying info from database.
        :param tbl_name: name of the table.
        :return: table as a pandas DataFrame with info as a tuple.
        """
        dataset_type = tbl_name.split('_')[0]

        if dataset_type == 'bloom':
            return self.__get_table_bloom(tbl_name)
        elif dataset_type == 'quandl':
            return self.__get_table_quandl(tbl_name)
        else:
            return None, None

    def get_info(self, tbl_name: str) -> Optional[Tuple[str, str, str, str, str]]:
        """
        Retrieves info for given table (asset name, type, ...)
        :param tbl_name: name of the table
        :return: info for table.
        """
        dataset_type = tbl_name.split('_')[0]

        if dataset_type == 'quandl':
            tbl_name = self.quandl_to_bloom_dict.get(tbl_name)[0]

            if tbl_name is None:
                return None

        result = self.con.execute('''SELECT * FROM bloom_info 
                                  WHERE table_name='{}' '''.format(tbl_name))
        info = result.fetchone()

        if info is not None:
            return info

        return None

    def get_assets_by_type(self, asset_type: str, asset_subtype: str = None) -> Optional[pd.DataFrame]:
        """
        Retrieves contract name and corrsponding table names based on given types.
        :param asset_type: contract type.
        :param asset_subtype: contract subtype.
        :return: a pandas Dataframe.
        """
        cursor = None

        assert(asset_type in ['Commodity', 'Equity', 'Interest Rates', 'Currency'])

        if asset_type is not None:
            if asset_subtype is not None:
                assert(asset_subtype in ['Developed', 'Emerging', 'Grains', 'Energy', \
                                      'Softs', 'Precious Metal', 'Meat', 'Metal'])

                cursor = self.con.execute('''SELECT table_name, contract_name
                                            FROM bloom_info
                                            WHERE type=? and subtype=?''', (asset_type, asset_subtype))

            else:
                cursor = self.con.execute('''SELECT table_name, contract_name
                                            FROM bloom_info
                                            WHERE type=? ''', (asset_type,))

            result = cursor.fetchall()

            if result is not None:
                df = pd.DataFrame(result, columns=['table_name', 'contract_name'])
                df['quandl_table_name'] = [self.bloom_to_qunadl_dict.get(tbl) for tbl in df['table_name']]
                df.set_index('table_name', inplace=True)

                return df

        return None

    def __get_table_bloom(self, tbl_name: str) -> Union[Tuple[pd.DataFrame, Tuple[str, str, str, str, str]], Tuple[None, None]]:
        """
        Retrieves bloom table from database.
        :param tbl_name: name of the table to retrieve
        :return: tuple consisting of table returned from database as pandas dataframe and table's info as a tuple
        containing [table name, contract name, symbol, type, subtype]
        """
        result = self.con.execute('''SELECT * FROM {}'''.format(tbl_name))
        table = result.fetchall()

        df = pd.DataFrame(table, columns=['Dates', 'PX_LAST', 'PX_OPEN', 'PX_HIGH',
                                          'PX_LOW'])
        df['Dates'] = pd.to_datetime(df['Dates'], format="%Y/%m/%d")
        df.set_index('Dates', drop=True, inplace=True)

        info = self.get_info(tbl_name)

        if tbl_name == 'bloom_sm1':
            df = df[df.index > datetime.datetime(1998, 1, 30)]

        elif tbl_name == 'bloom_xp1':
            df = df[df.index > datetime.datetime(2001, 1, 2)]

        elif tbl_name == 'bloom_er1':
            df = df[df.index > datetime.datetime(1999, 1, 11)]

        elif tbl_name == 'bloom_ih1':
            df = df[df.index > datetime.datetime(2005, 10, 10)]

        elif tbl_name == 'bloom_si1':
            df = df[df.index > datetime.datetime(1986, 12, 2)]

        elif tbl_name == 'bloom_ec1':
            df = df[df.index > datetime.datetime(1999, 1, 4)]

        elif tbl_name == 'bloom_nq1':
            df = df[df.index > datetime.datetime(2000, 5, 22)]

        elif tbl_name == 'bloom_ub1':
            df = df[df.index > datetime.datetime(2005, 9, 9)]

        elif tbl_name == 'bloom_xb1':
            df = df[df.index > datetime.datetime(2006, 4, 21)]

        elif tbl_name == 'bloom_nv1':
            df = df[df.index > datetime.datetime(2001, 8, 10)]

        elif tbl_name == 'bloom_sf1':
            df = df[df.index > datetime.datetime(1986, 7, 18)]

        elif tbl_name == 'bloom_fb1':
            df = df[df.index > datetime.datetime(1995, 6, 13)]

        elif tbl_name == 'bloom_g_1':
            df = df[df.index > datetime.datetime(1995, 6, 13)]

        elif tbl_name == 'bloom_ff1':
            return None, None

        elif tbl_name == 'bloom_ed1':
            return None, None

        elif tbl_name == 'bloom_l_1':
            return None, None

        elif tbl_name == 'bloom_rf1':
            return None, None

        elif tbl_name == 'bloom_lx1':
            return None, None

        elif tbl_name == 'bloom_la1':
            return None, None

        elif tbl_name == 'bloom_tr1':
            return None, None

        elif tbl_name == 'bloom_rl1':
            return None, None

        elif tbl_name == 'bloom_ll1':
            return None, None

        elif tbl_name == 'bloom_tm1':
            return None, None

        elif tbl_name == 'bloom_da1':
            return None, None

        df.drop_duplicates(inplace=True)

        return df, info

    def __get_table_quandl(self, tbl_name: str) -> Union[Tuple[pd.DataFrame, Tuple[str, str, str, str, str]], Tuple[None, None]]:
        """
        Retrieves quandl table from database.
        :param tbl_name: name of the table to retrieve
        :return: tuple consisting of table returned from database as pandas dataframe and table's info as a tuple
        containing [table name, contract name, symbol, type, subtype]
        """
        result = self.con.execute('''SELECT * FROM {}'''.format(tbl_name))
        table = result.fetchall()

        if table is None:
            return None, None

        df = pd.DataFrame(table, columns=['Date', 'Open', 'High', 'Low', 'Settle', 'Volume', 'Prev_day_OI'])

        # FIXME first rows should be dropped
        df = df.drop([0])

        df['Date'] = pd.to_datetime(df['Date'], format="%Y/%m/%d")
        df.rename(index=str, columns={'Date': 'Dates', 'Settle': 'PX_LAST', 'Open': 'PX_OPEN', 'High': 'PX_HIGH',
                                      'Low': 'PX_LOW'}, inplace=True)
        df.drop(['Volume', 'Prev_day_OI'], axis=1, inplace=True)
        df.set_index('Dates', drop=True, inplace=True)

        analogous_bloom = self.quandl_to_bloom_dict.get(tbl_name)

        info = None

        if analogous_bloom is not None:
            info = self.get_info(analogous_bloom[0])

        return df, info

    def __get_table_join(self, tbl_name_1: str, tbl_name_2: str) -> Optional[pd.DataFrame]:
        """
        Retrieves inner join of two tables from database.
        :param tbl_name_1: name of the left table in inner join.
        :param tbl_name_2: name of the right table in inner join.
        :return: join of tables as a pandas dataframe.

        """
        result = self.con.execute('''SELECT * FROM {} AS b, {} AS q
                                    WHERE b.Dates = q.Date '''.format(tbl_name_1, tbl_name_2))
        table = result.fetchall()

        df = pd.DataFrame(table, columns=['Dates', 'PX_LAST', 'PX_OPEN', 'PX_HIGH',
                                          'PX_LOW', 'Date', 'Open', 'High', 'Low',
                                          'Settle', 'Volume', 'Prev_OI'])

        if df.shape[0] == 0:
            return None

        df['Dates'] = pd.to_datetime(df['Dates'], format="%Y/%m/%d")
        df.drop('Date', inplace=True, axis=1)
        df.set_index('Dates', drop=False, inplace=True)

        return df

    def get_quandl_bloom_intersect(self) -> List[Tuple[pd.DataFrame, str, str]]:
        """
        Retrieves intersection of datasets in quandl & bloom.

        :return: list of pandas dataframes, each being a join of intersection of bloom & quandl datasets.
        """
        df_ret = []

        for tbl_quandl, tbl_bloom_list in self.quandl_to_bloom_dict.items():
            for tbl_bloom in tbl_bloom_list:
                joined_df = self.__get_table_join(tbl_quandl, tbl_bloom)

                if joined_df is not None:
                    df_ret.append((joined_df, tbl_quandl, tbl_bloom))

        return df_ret
